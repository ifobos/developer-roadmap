#!/usr/bin/env python3
"""
Script para verificar que los archivos Excel generados contengan los datos correctos
"""

import os
import pandas as pd
from openpyxl import load_workbook

def verify_excel_file(excel_path, original_csv_path):
    """
    Verifica que un archivo Excel contenga los datos correctos del CSV original
    """
    print(f"\nVerificando: {os.path.basename(excel_path)}")
    
    try:
        # Leer Excel
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Obtener nombres de columnas (fila 2)
        column_names = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col).value
            if cell_value:
                column_names.append(str(cell_value).strip())
            else:
                break
        
        print(f"  Columnas en Excel: {column_names}")
        
        # Contar filas de datos (desde fila 4)
        data_rows = 0
        for row in range(4, ws.max_row + 1):
            if ws.cell(row=row, column=1).value:  # Si hay valor en primera columna
                data_rows += 1
            else:
                break
        
        print(f"  Filas de datos en Excel: {data_rows}")
        
        # Leer CSV original para comparar
        if os.path.exists(original_csv_path):
            df_csv = pd.read_csv(original_csv_path)
            print(f"  Filas en CSV original: {len(df_csv)}")
            
            # Verificar que coincidan las filas
            if data_rows == len(df_csv):
                print(f"  ✓ Número de filas coincide")
            else:
                print(f"  ✗ Número de filas NO coincide")
            
            # Mostrar muestra de datos del Excel
            print(f"  Muestra de datos del Excel:")
            for row in range(4, min(7, ws.max_row + 1)):  # Mostrar primeras 3 filas de datos
                row_data = []
                for col in range(1, len(column_names) + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else "")
                print(f"    Fila {row-3}: {dict(zip(column_names, row_data))}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ✗ Error verificando archivo: {e}")
        return False

def main():
    """
    Función principal
    """
    # Usar ruta relativa al directorio padre desde InterviewTemplates
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_dir = os.path.join(base_dir, "Interviews", "generated")
    csv_dir = os.path.join(base_dir, "csv_output")
    
    # Obtener lista de archivos Excel generados
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('_interview_template.xlsx')]
    
    if not excel_files:
        print(f"No se encontraron archivos Excel en {excel_dir}")
        return
    
    print(f"Verificando {len(excel_files)} archivos Excel generados...")
    
    # Verificar algunos archivos como muestra
    sample_files = sorted(excel_files)[:5]  # Primeros 5 archivos
    
    successful = 0
    failed = 0
    
    for excel_file in sample_files:
        excel_path = os.path.join(excel_dir, excel_file)
        
        # Determinar el archivo CSV correspondiente
        roadmap_name = excel_file.replace('_interview_template.xlsx', '')
        csv_file = f"{roadmap_name}_nodes.csv"
        csv_path = os.path.join(csv_dir, csv_file)
        
        if verify_excel_file(excel_path, csv_path):
            successful += 1
        else:
            failed += 1
    
    print(f"\n=== RESUMEN DE VERIFICACIÓN ===")
    print(f"Archivos verificados exitosamente: {successful}")
    print(f"Archivos con errores: {failed}")
    print(f"Total verificados: {len(sample_files)}")
    print(f"Total archivos Excel generados: {len(excel_files)}")

if __name__ == "__main__":
    main()