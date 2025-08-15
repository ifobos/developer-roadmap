#!/usr/bin/env python3
"""
Script para generar archivos Excel individuales basados en un template Excel
y datos de archivos CSV del directorio csv_output.

El script:
1. Lee el template Excel (InterviewTemplate.xlsx)
2. Para cada archivo CSV en csv_output/
3. Crea un nuevo archivo Excel basado en el template
4. Mapea las columnas: roadmap, label, content_file_path
5. Copia las fórmulas de la fila 3 del template
6. Guarda el archivo Excel en el directorio Interviews/
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.comments import Comment
import shutil
from pathlib import Path

def read_content_file(content_file_path, base_dir):
    """
    Lee el contenido de un archivo referenciado en content_file_path
    """
    if pd.isna(content_file_path) or content_file_path == '':
        return ""
    
    # Construir la ruta completa del archivo
    full_path = os.path.join(base_dir, content_file_path)
    
    try:
        with open(full_path, 'r', encoding='utf-8') as file:
            content = file.read().strip()
            return content
    except FileNotFoundError:
        print(f"    Advertencia: Archivo no encontrado: {full_path}")
        return ""
    except Exception as e:
        print(f"    Error leyendo archivo {full_path}: {e}")
        return ""

def read_template_structure(template_path):
    """
    Lee la estructura del template Excel para entender:
    - Nombres de columnas (fila 2)
    - Fórmulas y formato (fila 3)
    """
    print(f"Leyendo template: {template_path}")
    
    # Cargar el workbook del template
    wb_template = load_workbook(template_path)
    ws_template = wb_template.active
    
    # Leer nombres de columnas de la fila 2
    column_names = []
    for col in range(1, ws_template.max_column + 1):
        cell_value = ws_template.cell(row=2, column=col).value
        if cell_value:
            column_names.append(str(cell_value).strip())
        else:
            break
    
    print(f"Columnas encontradas en template: {column_names}")
    
    # Leer fórmulas y estilos de la fila 3 (fila de ejemplo/formato)
    template_row_data = []
    for col in range(1, len(column_names) + 1):
        cell = ws_template.cell(row=3, column=col)
        template_row_data.append({
            'value': cell.value,
            'formula': cell.formula if hasattr(cell, 'formula') else None,
            'font': cell.font,
            'alignment': cell.alignment,
            'fill': cell.fill,
            'number_format': cell.number_format
        })
    
    return column_names, template_row_data, wb_template

def process_csv_file(csv_path, template_path, output_dir):
    """
    Procesa un archivo CSV y genera el archivo Excel correspondiente
    """
    csv_filename = os.path.basename(csv_path)
    roadmap_name = csv_filename.replace('_nodes.csv', '')
    
    print(f"\nProcesando: {csv_filename} -> {roadmap_name}")
    
    # Directorio base para resolver rutas de archivos de contenido
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Leer CSV
    try:
        df_csv = pd.read_csv(csv_path)
        print(f"  CSV leído: {len(df_csv)} filas")
    except Exception as e:
        print(f"  Error leyendo CSV: {e}")
        return False
    
    # Verificar que el CSV tenga las columnas necesarias
    required_columns = ['roadmap', 'label']
    missing_columns = [col for col in required_columns if col not in df_csv.columns]
    
    if missing_columns:
        print(f"  Advertencia: Columnas faltantes en CSV: {missing_columns}")
    
    # Leer estructura del template
    try:
        column_names, template_row_data, wb_template = read_template_structure(template_path)
    except Exception as e:
        print(f"  Error leyendo template: {e}")
        return False
    
    # Crear nuevo workbook copiando el template
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_template = wb_template.active
    
    # Copiar las primeras 3 filas del template (encabezados y formato)
    for row in range(1, 4):
        for col in range(1, len(column_names) + 1):
            source_cell = ws_template.cell(row=row, column=col)
            target_cell = ws_new.cell(row=row, column=col)
            
            # Copiar valor
            target_cell.value = source_cell.value
            
            # Copiar estilo
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
            
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            target_cell.number_format = source_cell.number_format
    
    # Mapear columnas CSV a columnas del template
    column_mapping = {}
    for i, col_name in enumerate(column_names):
        if col_name in df_csv.columns:
            column_mapping[i] = col_name
        else:
            column_mapping[i] = None  # Usar valor del template
    
    print(f"  Mapeo de columnas: {column_mapping}")
    
    # Filtrar filas que cumplan con los criterios: label y content_file_path no vacíos
    if 'content_file_path' in df_csv.columns:
        filtered_df = df_csv[
            (df_csv['label'].notna()) & 
            (df_csv['label'] != '') & 
            (df_csv['content_file_path'].notna()) & 
            (df_csv['content_file_path'] != '')
        ].reset_index(drop=True)
        print(f"  Filas filtradas: {len(filtered_df)} de {len(df_csv)} (con label y content_file_path válidos)")
    else:
        filtered_df = df_csv[
            (df_csv['label'].notna()) & 
            (df_csv['label'] != '')
        ].reset_index(drop=True)
        print(f"  Filas filtradas: {len(filtered_df)} de {len(df_csv)} (solo con label válido)")
    
    # Process each row of data
    for csv_row_idx, csv_row in filtered_df.iterrows():
        excel_row = csv_row_idx + 4  # Empezar desde la fila 4
        content_text = None  # Almacenar contenido para agregar como nota
        label_cell = None  # Referencia a la celda LABEL
        
        for col_idx, template_col_name in enumerate(column_names):
            excel_col = col_idx + 1
            cell = ws_new.cell(row=excel_row, column=excel_col)
            
            # Determinar el valor de la celda
            if template_col_name in df_csv.columns:
                # Usar valor del CSV
                csv_value = csv_row[template_col_name]
                # Convertir NaN a string vacío
                if pd.isna(csv_value):
                    cell.value = ""
                else:
                    cell.value = str(csv_value)
                    
                # Guardar referencia a la celda LABEL y leer contenido para nota
                if template_col_name.upper() == 'LABEL':
                    label_cell = cell
                    # Leer contenido del archivo para agregar como nota
                    if 'content_file_path' in df_csv.columns:
                        content_file_path = csv_row.get('content_file_path', '')
                        content_text = read_content_file(content_file_path, base_dir)
            else:
                # Usar valor del template (fila 3)
                template_cell_data = template_row_data[col_idx]
                
                if template_cell_data['value'] and str(template_cell_data['value']).startswith('='):
                    # Si hay fórmula, traducirla para la nueva fila
                    try:
                        formula = str(template_cell_data['value'])
                        # Reemplazar referencias de fila 3 por la fila actual
                        new_formula = formula.replace('3', str(excel_row))
                        cell.value = new_formula
                    except:
                        cell.value = template_cell_data['value']
                else:
                    cell.value = template_cell_data['value']
            
            # Copiar estilo de la fila 3 del template
            template_cell_data = template_row_data[col_idx]
            
            if template_cell_data['font']:
                cell.font = Font(
                    name=template_cell_data['font'].name,
                    size=template_cell_data['font'].size,
                    bold=template_cell_data['font'].bold,
                    italic=template_cell_data['font'].italic,
                    color=template_cell_data['font'].color
                )
            
            if template_cell_data['alignment']:
                cell.alignment = Alignment(
                    horizontal=template_cell_data['alignment'].horizontal,
                    vertical=template_cell_data['alignment'].vertical,
                    wrap_text=template_cell_data['alignment'].wrap_text
                )
            
            if template_cell_data['fill']:
                cell.fill = PatternFill(
                    fill_type=template_cell_data['fill'].fill_type,
                    start_color=template_cell_data['fill'].start_color,
                    end_color=template_cell_data['fill'].end_color
                )
            
            cell.number_format = template_cell_data['number_format']
        
        # Agregar contenido como nota a la celda LABEL si tenemos contenido y celda LABEL
        if content_text and label_cell and content_text.strip():
            comment = Comment(content_text, "Sistema")
            comment.width = 400
            comment.height = 300
            label_cell.comment = comment
    
    # Ajustar ancho de columnas
    for col in ws_new.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_new.column_dimensions[column].width = adjusted_width
    
    # Guardar archivo Excel
    output_filename = f"{roadmap_name}_interview_template.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    try:
        wb_new.save(output_path)
        print(f"  ✓ Archivo guardado: {output_path}")
        return True
    except Exception as e:
        print(f"  ✗ Error guardando archivo: {e}")
        return False
    finally:
        wb_new.close()
        wb_template.close()

def main():
    """
    Función principal
    """
    # Rutas (relativas al directorio padre)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, "Interviews", "InterviewTemplate.xlsx")
    csv_dir = os.path.join(base_dir, "csv_output")
    output_dir = os.path.join(base_dir, "Interviews", "generated")
    
    # Verificar que existe el template
    if not os.path.exists(template_path):
        print(f"Error: No se encuentra el template en {template_path}")
        return
    
    # Crear directorio de salida si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Obtener lista de archivos CSV
    csv_files = [f for f in os.listdir(csv_dir) if f.endswith('_nodes.csv')]
    
    if not csv_files:
        print(f"No se encontraron archivos CSV en {csv_dir}")
        return
    
    print(f"Encontrados {len(csv_files)} archivos CSV para procesar")
    print(f"Template: {template_path}")
    print(f"Directorio de salida: {output_dir}")
    
    # Procesar cada archivo CSV
    successful = 0
    failed = 0
    
    for csv_file in sorted(csv_files):
        csv_path = os.path.join(csv_dir, csv_file)
        
        if process_csv_file(csv_path, template_path, output_dir):
            successful += 1
        else:
            failed += 1
    
    print(f"\n=== RESUMEN ===")
    print(f"Archivos procesados exitosamente: {successful}")
    print(f"Archivos con errores: {failed}")
    print(f"Total: {len(csv_files)}")
    
    if successful > 0:
        print(f"\nArchivos Excel generados en: {output_dir}")

if __name__ == "__main__":
    main()